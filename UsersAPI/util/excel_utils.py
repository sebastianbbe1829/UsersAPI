import io
from datetime import datetime
from typing import Any

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def _obtener_datos_usuario(current_user: Any | None) -> tuple[str, str]:
    """
    Obtiene de forma segura el nombre y documento del usuario
    que genera el reporte.

    Soporta objetos como:
    - UserDB
    - UserTenantDB
    - GlobalUserDB
    - Objetos con relaciones de usuario
    """

    if current_user is None:
        return (
            "Usuario no identificado",
            "N/A",
        )

    # ==========================================================
    # CASO 1: EL OBJETO YA TIENE NAME Y DNI
    # ==========================================================

    nombre = getattr(current_user, "name", None)
    dni = getattr(current_user, "dni", None)

    if nombre:
        return (
            str(nombre),
            str(dni) if dni else "N/A",
        )

    # ==========================================================
    # CASO 2: USER TENANT CON RELACIÓN "user"
    # ==========================================================

    usuario = getattr(current_user, "user", None)

    if usuario is not None:
        nombre = getattr(usuario, "name", None)
        dni = getattr(usuario, "dni", None)

        if nombre:
            return (
                str(nombre),
                str(dni) if dni else "N/A",
            )

    # ==========================================================
    # CASO 3: RELACIÓN CON APP_USER
    # ==========================================================

    app_user = getattr(current_user, "app_user", None)

    if app_user is not None:
        nombre = getattr(app_user, "name", None)
        dni = getattr(app_user, "dni", None)

        if nombre:
            return (
                str(nombre),
                str(dni) if dni else "N/A",
            )

    # ==========================================================
    # CASO 4: GLOBAL USER
    # ==========================================================

    email = getattr(current_user, "email", None)

    if email:
        return (
            str(email),
            str(dni) if dni else "N/A",
        )

    # ==========================================================
    # FALLBACK FINAL
    # ==========================================================

    return (
        "Usuario no identificado",
        str(dni) if dni else "N/A",
    )


def export_to_excel(
    data: list[dict],
    filename: str,
    current_user: Any | None = None,
):
    """
    Genera un reporte de usuarios en Excel con formato visual,
    información del usuario que genera el reporte y resumen.
    """

    # ==========================================================
    # CREAR LIBRO
    # ==========================================================

    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    # ==========================================================
    # COLORES
    # ==========================================================

    COLOR_PRINCIPAL = "1F4E78"
    COLOR_SECUNDARIO = "D9EAF7"
    COLOR_TITULO = "17365D"
    COLOR_BLANCO = "FFFFFF"
    COLOR_GRIS = "F2F2F2"
    COLOR_GRIS_CLARO = "F8FAFC"
    COLOR_BORDE = "D9E1F2"
    COLOR_ACTIVO = "E2F0D9"
    COLOR_INACTIVO = "FCE4D6"

    # ==========================================================
    # BORDES
    # ==========================================================

    borde_suave = Border(
        left=Side(
            style="thin",
            color=COLOR_BORDE,
        ),
        right=Side(
            style="thin",
            color=COLOR_BORDE,
        ),
        top=Side(
            style="thin",
            color=COLOR_BORDE,
        ),
        bottom=Side(
            style="thin",
            color=COLOR_BORDE,
        ),
    )

    # ==========================================================
    # TÍTULO
    # ==========================================================

    ws.merge_cells("A1:E1")

    ws["A1"] = "REPORTE DE USUARIOS"

    ws["A1"].font = Font(
        bold=True,
        size=20,
        color=COLOR_BLANCO,
    )

    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=COLOR_TITULO,
    )

    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws.row_dimensions[1].height = 35

    # ==========================================================
    # INFORMACIÓN DEL REPORTE
    # ==========================================================

    fecha_generacion = datetime.now().strftime(
        "%d/%m/%Y %H:%M:%S"
    )

    nombre_usuario, dni_usuario = (
        _obtener_datos_usuario(current_user)
    )

    informacion = [
        (
            "Generado por",
            nombre_usuario,
        ),
        (
            "DNI",
            dni_usuario,
        ),
        (
            "Fecha de generación",
            fecha_generacion,
        ),
    ]

    fila = 3

    for etiqueta, valor in informacion:

        celda_etiqueta = ws.cell(
            row=fila,
            column=1,
            value=etiqueta,
        )

        celda_valor = ws.cell(
            row=fila,
            column=2,
            value=valor,
        )

        celda_etiqueta.font = Font(
            bold=True,
            color=COLOR_TITULO,
        )

        celda_etiqueta.fill = PatternFill(
            "solid",
            fgColor=COLOR_SECUNDARIO,
        )

        celda_etiqueta.border = borde_suave
        celda_valor.border = borde_suave

        fila += 1

    # ==========================================================
    # RESUMEN
    # ==========================================================

    total_usuarios = len(data)

    activos = sum(
        1
        for usuario in data
        if usuario.get("Estado") == "Activo"
    )

    inactivos = sum(
        1
        for usuario in data
        if usuario.get("Estado") == "Inactivo"
    )

    fila_resumen = 7

    celda_resumen = ws.cell(
        row=fila_resumen,
        column=1,
        value="RESUMEN",
    )

    celda_resumen.font = Font(
        bold=True,
        color=COLOR_BLANCO,
    )

    celda_resumen.fill = PatternFill(
        "solid",
        fgColor=COLOR_PRINCIPAL,
    )

    celda_resumen.alignment = Alignment(
        horizontal="center",
    )

    celda_resumen.border = borde_suave

    resumen = [
        (
            "Total usuarios",
            total_usuarios,
        ),
        (
            "Usuarios activos",
            activos,
        ),
        (
            "Usuarios inactivos",
            inactivos,
        ),
    ]

    for columna, (
        etiqueta,
        valor,
    ) in enumerate(
        resumen,
        start=2,
    ):

        celda = ws.cell(
            row=fila_resumen,
            column=columna,
            value=f"{etiqueta}: {valor}",
        )

        celda.font = Font(
            bold=True,
            color=COLOR_TITULO,
        )

        celda.fill = PatternFill(
            "solid",
            fgColor=COLOR_GRIS,
        )

        celda.alignment = Alignment(
            horizontal="center",
        )

        celda.border = borde_suave

    # ==========================================================
    # ENCABEZADOS
    # ==========================================================

    fila_inicio_tabla = 10

    columnas = [
        "DNI",
        "Nombre",
        "Email",
        "Teléfono",
        "Estado",
    ]

    for columna, encabezado in enumerate(
        columnas,
        start=1,
    ):

        celda = ws.cell(
            row=fila_inicio_tabla,
            column=columna,
            value=encabezado,
        )

        celda.font = Font(
            bold=True,
            color=COLOR_BLANCO,
        )

        celda.fill = PatternFill(
            "solid",
            fgColor=COLOR_PRINCIPAL,
        )

        celda.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        celda.border = borde_suave

    ws.row_dimensions[
        fila_inicio_tabla
    ].height = 25

    # ==========================================================
    # DATOS
    # ==========================================================

    fila_actual = fila_inicio_tabla + 1

    for indice, usuario in enumerate(data):

        for columna, encabezado in enumerate(
            columnas,
            start=1,
        ):

            valor = usuario.get(
                encabezado,
                "",
            )

            celda = ws.cell(
                row=fila_actual,
                column=columna,
                value=valor,
            )

            celda.border = borde_suave

            celda.alignment = Alignment(
                vertical="center",
            )

            if indice % 2 == 1:

                celda.fill = PatternFill(
                    "solid",
                    fgColor=COLOR_GRIS_CLARO,
                )

        # ======================================================
        # DNI
        # ======================================================

        ws.cell(
            row=fila_actual,
            column=1,
        ).alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        # ======================================================
        # TELÉFONO
        # ======================================================

        ws.cell(
            row=fila_actual,
            column=4,
        ).alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        # ======================================================
        # ESTADO
        # ======================================================

        estado = usuario.get("Estado")

        celda_estado = ws.cell(
            row=fila_actual,
            column=5,
        )

        celda_estado.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        celda_estado.font = Font(
            bold=True,
        )

        if estado == "Activo":

            celda_estado.fill = PatternFill(
                "solid",
                fgColor=COLOR_ACTIVO,
            )

        elif estado == "Inactivo":

            celda_estado.fill = PatternFill(
                "solid",
                fgColor=COLOR_INACTIVO,
            )

        fila_actual += 1

    # ==========================================================
    # FILTRO DE EXCEL
    # ==========================================================

    ws.auto_filter.ref = (
        f"A{fila_inicio_tabla}:E{fila_actual - 1}"
        if total_usuarios > 0
        else (
            f"A{fila_inicio_tabla}:"
            f"E{fila_inicio_tabla}"
        )
    )

    # ==========================================================
    # ANCHOS DE COLUMNAS
    # ==========================================================

    anchos = {
        "A": 16,
        "B": 30,
        "C": 35,
        "D": 18,
        "E": 16,
    }

    for columna, ancho in anchos.items():

        ws.column_dimensions[
            columna
        ].width = ancho

    # ==========================================================
    # CONFIGURACIÓN DE LA HOJA
    # ==========================================================

    ws.freeze_panes = "A11"

    ws.sheet_view.showGridLines = False

    # ==========================================================
    # ALTURA DE FILAS
    # ==========================================================

    for fila_dato in range(
        fila_inicio_tabla + 1,
        fila_actual,
    ):

        ws.row_dimensions[
            fila_dato
        ].height = 22

    # ==========================================================
    # GENERAR ARCHIVO
    # ==========================================================

    output = io.BytesIO()

    wb.save(output)

    output.seek(0)

    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                f'attachment; filename="{filename}"'
            )
        },
    )