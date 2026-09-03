import io
from datetime import datetime
from typing import Any

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def _obtener_usuario(current_user: Any | None) -> tuple[str, str]:
    if current_user is None:
        return "Usuario no identificado", "N/A"
    usuario = getattr(current_user, "user", None)
    nombre = getattr(usuario, "name", None) or getattr(current_user, "name", None)
    dni = getattr(usuario, "dni", None) or getattr(current_user, "dni", None)
    if nombre:
        return str(nombre), str(dni) if dni else "N/A"
    email = getattr(current_user, "email", None)
    if email:
        return str(email), str(dni) if dni else "N/A"
    return "Usuario no identificado", str(dni) if dni else "N/A"


def export_extinguishers_to_excel(data: list[dict], current_user: Any | None = None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Extintores"

    principal, titulo, secundario = "1F4E78", "17365D", "D9EAF7"
    blanco, gris, gris_claro = "FFFFFF", "F2F2F2", "F8FAFC"
    borde_color, verde, amarillo, rojo = "D9E1F2", "E2F0D9", "FFF2CC", "FCE4D6"
    borde = Border(
        left=Side(style="thin", color=borde_color),
        right=Side(style="thin", color=borde_color),
        top=Side(style="thin", color=borde_color),
        bottom=Side(style="thin", color=borde_color),
    )

    columnas = [
        "Código", "Tipo", "Capacidad", "Ubicación", "Estado", "Stock",
        "Última recarga", "Próxima recarga", "Última revisión",
        "Resultado última revisión", "Revisiones desde hidrostática",
        "Última prueba hidrostática", "Próxima prueba hidrostática",
        "Hidrostática requerida",
    ]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columnas))
    ws.cell(1, 1, "REPORTE DE EXTINTORES")
    ws.cell(1, 1).font = Font(bold=True, size=20, color=blanco)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=titulo)
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    nombre, dni = _obtener_usuario(current_user)
    metadata = [
        ("Generado por", nombre),
        ("DNI", dni),
        ("Fecha de generación", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
    ]
    for fila, (etiqueta, valor) in enumerate(metadata, start=3):
        ws.cell(fila, 1, etiqueta).font = Font(bold=True, color=titulo)
        ws.cell(fila, 1).fill = PatternFill("solid", fgColor=secundario)
        ws.cell(fila, 1).border = borde
        ws.cell(fila, 2, valor).border = borde

    total = len(data)
    activos = sum(1 for x in data if x.get("Estado") == "Activo")
    inactivos = total - activos
    requiere = sum(1 for x in data if x.get("Hidrostática requerida") == "Sí")
    ws.cell(7, 1, "RESUMEN").font = Font(bold=True, color=blanco)
    ws.cell(7, 1).fill = PatternFill("solid", fgColor=principal)
    ws.cell(7, 1).alignment = Alignment(horizontal="center")
    ws.cell(7, 1).border = borde
    resumen = [
        f"Total extintores: {total}",
        f"Activos: {activos}",
        f"Inactivos: {inactivos}",
        f"Hidrostática requerida: {requiere}",
    ]
    for columna, texto in enumerate(resumen, start=2):
        celda = ws.cell(7, columna, texto)
        celda.font = Font(bold=True, color=titulo)
        celda.fill = PatternFill("solid", fgColor=gris)
        celda.alignment = Alignment(horizontal="center")
        celda.border = borde

    fila_encabezados = 10
    for columna, encabezado in enumerate(columnas, start=1):
        celda = ws.cell(fila_encabezados, columna, encabezado)
        celda.font = Font(bold=True, color=blanco)
        celda.fill = PatternFill("solid", fgColor=principal)
        celda.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        celda.border = borde
    ws.row_dimensions[fila_encabezados].height = 35

    for indice, item in enumerate(data):
        fila = fila_encabezados + 1 + indice
        for columna, encabezado in enumerate(columnas, start=1):
            celda = ws.cell(fila, columna, item.get(encabezado, ""))
            celda.border = borde
            celda.alignment = Alignment(vertical="center", wrap_text=True)
            if indice % 2:
                celda.fill = PatternFill("solid", fgColor=gris_claro)
        estado = ws.cell(fila, 5)
        estado.font = Font(bold=True)
        estado.alignment = Alignment(horizontal="center", vertical="center")
        estado.fill = PatternFill(
            "solid",
            fgColor=verde if item.get("Estado") == "Activo" else rojo,
        )
        hidrost = ws.cell(fila, 14)
        hidrost.alignment = Alignment(horizontal="center", vertical="center")
        if item.get("Hidrostática requerida") == "Sí":
            hidrost.fill = PatternFill("solid", fgColor=amarillo)
            hidrost.font = Font(bold=True)

    ultima_fila = fila_encabezados + max(total, 1)
    ws.auto_filter.ref = f"A{fila_encabezados}:N{ultima_fila}"
    ws.freeze_panes = "A11"
    ws.sheet_view.showGridLines = False
    anchos = [18, 28, 15, 28, 14, 12, 18, 18, 18, 25, 24, 22, 22, 24]
    for indice, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[chr(64 + indice)].width = ancho
    for fila in range(fila_encabezados + 1, ultima_fila + 1):
        ws.row_dimensions[fila].height = 24

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="extintores.xlsx"',
        },
    )
