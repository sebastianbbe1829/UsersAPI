import base64
import io
import os
from datetime import date, datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import Column, Date, DateTime, Integer, MetaData, String, Table, and_, func, select, update
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.orm import Session

from ..logging_config import logger
from ..models import ExtinguisherDB, ExtinguisherTypeDB, RoleDB, TenantDB, UserTenantDB, UserTenantRoleDB
from ..util.email_utils import send_email

ADMIN_ROLE_CODES = ("ADMIN",)
REPORT_TIMEZONE = os.getenv("JOB_TIMEZONE", "America/Bogota")
_NOTIFICATION_METADATA = MetaData()
NOTIFICATION_LOG_TABLE = Table(
    "extinguisher_recharge_notification_log",
    _NOTIFICATION_METADATA,
    Column("notification_date", Date, nullable=False),
    Column("tenant_id", Integer, nullable=False),
    Column("recipient", String(320), nullable=False),
    Column("status", String(20), nullable=False),
    Column("sent_at", DateTime(timezone=True), nullable=True),
    schema="users_api",
)


class ExtinguisherRechargeNotificationService:
    """Envía el aviso diario de extintores vencidos o con recarga para hoy."""

    def __init__(self, db: Session):
        self.db = db

    def run(self, notification_date: date | None = None) -> dict:
        target_date = notification_date or datetime.now(ZoneInfo(REPORT_TIMEZONE)).date()
        rows = self.db.execute(
            select(
                ExtinguisherDB,
                ExtinguisherTypeDB.name.label("type_name"),
                TenantDB.name.label("tenant_name"),
                TenantDB.slug.label("tenant_slug"),
            )
            .join(ExtinguisherTypeDB, ExtinguisherTypeDB.id == ExtinguisherDB.extinguisher_type_id)
            .join(TenantDB, TenantDB.id == ExtinguisherDB.tenant_id)
            .where(
                and_(
                    ExtinguisherDB.active.is_(True),
                    ExtinguisherDB.next_recharge_date <= target_date,
                    TenantDB.status == 1,
                )
            )
            .order_by(
                TenantDB.id,
                ExtinguisherDB.next_recharge_date.asc(),
                ExtinguisherDB.id.asc(),
            )
        ).all()

        by_tenant: dict[int, dict] = {}
        for extinguisher, type_name, tenant_name, tenant_slug in rows:
            bucket = by_tenant.setdefault(
                extinguisher.tenant_id,
                {
                    "tenant_name": tenant_name,
                    "tenant_slug": tenant_slug,
                    "extinguishers": [],
                },
            )
            days_overdue = (target_date - extinguisher.next_recharge_date).days
            bucket["extinguishers"].append(
                {
                    "code": extinguisher.code,
                    "type_name": type_name,
                    "capacity": extinguisher.capacity,
                    "location": extinguisher.location,
                    "last_recharge_date": extinguisher.last_recharge_date,
                    "next_recharge_date": extinguisher.next_recharge_date,
                    "days_overdue": days_overdue,
                    "status": "VENCE HOY" if days_overdue == 0 else "VENCIDO",
                }
            )

        sent = skipped = errors = 0
        for tenant_id, data in by_tenant.items():
            recipients = self._get_admin_recipients(tenant_id)
            if not recipients:
                logger.warning(
                    "No active ADMIN email found for tenant_id=%s; recharge notification skipped",
                    tenant_id,
                )
                skipped += 1
                continue

            attachment = self._build_excel_attachment(data["tenant_name"], target_date, data["extinguishers"])
            message = self._build_message(data["tenant_name"], target_date, data["extinguishers"])
            overdue_count = sum(1 for item in data["extinguishers"] if item["days_overdue"] > 0)
            today_count = len(data["extinguishers"]) - overdue_count
            subject = (
                f"Alerta: {len(data['extinguishers'])} extintor(es) vencido(s) "
                f"o con recarga hoy - {data['tenant_name']}"
            )
            message = (
                message.replace("{TOTAL}", str(len(data["extinguishers"])))
                .replace("{VENCIDOS}", str(overdue_count))
                .replace("{HOY}", str(today_count))
            )

            for recipient in recipients:
                recipient_key = recipient.strip().lower()
                if self._already_sent(target_date, tenant_id, recipient_key):
                    continue
                self._mark_pending(target_date, tenant_id, recipient_key)
                try:
                    send_email(
                        recipient=recipient,
                        subject=subject,
                        message=message,
                        tenant_slug=data["tenant_slug"],
                        tenant_name=data["tenant_name"],
                        template="default",
                        attachments=[attachment],
                    )
                    self._mark_sent(target_date, tenant_id, recipient_key)
                    sent += 1
                except Exception:
                    self._clear_pending(target_date, tenant_id, recipient_key)
                    errors += 1
                    logger.exception(
                        "Error sending recharge notification to %s for tenant_id=%s",
                        recipient,
                        tenant_id,
                    )

        result = {
            "date": target_date.isoformat(),
            "tenants_with_expired_recharges": len(by_tenant),
            "extinguishers": len(rows),
            "emails_sent": sent,
            "tenants_without_admin_email": skipped,
            "email_errors": errors,
        }
        logger.info("Daily extinguisher recharge notification finished: %s", result)
        return result

    def _build_excel_attachment(self, tenant_name: str, target_date: date, extinguishers: list[dict]) -> dict[str, str]:
        wb = Workbook()
        ws = wb.active
        ws.title = "Recargas pendientes"
        headers = ["Código", "Tipo", "Capacidad", "Ubicación", "Última recarga", "Próxima recarga", "Días vencido", "Estado"]
        principal, dark, white, gray = "1F4E78", "17365D", "FFFFFF", "F2F2F2"
        border = Border(
            left=Side(style="thin", color="D9E1F2"),
            right=Side(style="thin", color="D9E1F2"),
            top=Side(style="thin", color="D9E1F2"),
            bottom=Side(style="thin", color="D9E1F2"),
        )
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(1, 1, f"REPORTE DE RECARGAS PENDIENTES - {tenant_name}")
        ws.cell(1, 1).font = Font(bold=True, size=16, color=white)
        ws.cell(1, 1).fill = PatternFill("solid", fgColor=dark)
        ws.cell(1, 1).alignment = Alignment(horizontal="center")
        ws.cell(2, 1, "Fecha del reporte")
        ws.cell(2, 2, target_date)
        ws.cell(2, 2).number_format = "dd/mm/yyyy"
        for col, header in enumerate(headers, 1):
            cell = ws.cell(4, col, header)
            cell.font = Font(bold=True, color=white)
            cell.fill = PatternFill("solid", fgColor=principal)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = border
        for row_idx, item in enumerate(extinguishers, 5):
            values = [item["code"], item["type_name"], item["capacity"], item["location"], item["last_recharge_date"], item["next_recharge_date"], item["days_overdue"], item["status"]]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row_idx, col, value)
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if row_idx % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=gray)
            for col in (5, 6):
                ws.cell(row_idx, col).number_format = "dd/mm/yyyy"
            ws.cell(row_idx, 7).alignment = Alignment(horizontal="center")
            ws.cell(row_idx, 8).font = Font(bold=True)
        ws.auto_filter.ref = f"A4:H{max(4, len(extinguishers) + 4)}"
        ws.freeze_panes = "A5"
        ws.sheet_view.showGridLines = False
        for col, width in enumerate((18, 28, 15, 30, 18, 18, 15, 15), 1):
            ws.column_dimensions[chr(64 + col)].width = width
        output = io.BytesIO()
        wb.save(output)
        encoded = base64.b64encode(output.getvalue()).decode("ascii")
        return {"name": f"extintores_recarga_{target_date.isoformat()}.xlsx", "content": encoded}

    @staticmethod
    def _build_message(tenant_name: str, target_date: date, extinguishers: list[dict]) -> str:
        overdue = sum(1 for item in extinguishers if item["days_overdue"] > 0)
        lines = [
            f"Buenos días, {tenant_name}.",
            "",
            (
                f"El reporte diario de {target_date.strftime('%d/%m/%Y')} identifica "
                "{TOTAL} extintor(es) con recarga pendiente."
            ),
            "",
            "- Vencidos: {VENCIDOS}",
            "- Vencen hoy: {HOY}",
            "",
            "Se adjunta el Excel con el detalle de los extintores, sus fechas de recarga y los días de vencimiento.",
            "",
            "Por favor, realiza la gestión correspondiente de recarga y actualiza la información en el sistema.",
            "",
            "Este es un mensaje automático.",
        ]
        return "\n".join(lines)

    def _already_sent(self, target_date, tenant_id, recipient):
        statement = select(NOTIFICATION_LOG_TABLE.c.status).where(
            and_(
                NOTIFICATION_LOG_TABLE.c.notification_date == target_date,
                NOTIFICATION_LOG_TABLE.c.tenant_id == tenant_id,
                NOTIFICATION_LOG_TABLE.c.recipient == recipient,
            )
        )
        return self.db.execute(statement).scalar_one_or_none() == "sent"

    def _mark_pending(self, target_date, tenant_id, recipient):
        statement = (
            pg_insert(NOTIFICATION_LOG_TABLE)
            .values(notification_date=target_date, tenant_id=tenant_id, recipient=recipient, status="pending")
            .on_conflict_do_update(
                index_elements=[
                    NOTIFICATION_LOG_TABLE.c.notification_date,
                    NOTIFICATION_LOG_TABLE.c.tenant_id,
                    NOTIFICATION_LOG_TABLE.c.recipient,
                ],
                set_={"status": "pending"},
                where=NOTIFICATION_LOG_TABLE.c.status != "sent",
            )
        )
        self.db.execute(statement)
        self.db.commit()

    def _clear_pending(self, target_date, tenant_id, recipient):
        statement = NOTIFICATION_LOG_TABLE.delete().where(
            and_(
                NOTIFICATION_LOG_TABLE.c.notification_date == target_date,
                NOTIFICATION_LOG_TABLE.c.tenant_id == tenant_id,
                NOTIFICATION_LOG_TABLE.c.recipient == recipient,
                NOTIFICATION_LOG_TABLE.c.status == "pending",
            )
        )
        self.db.execute(statement)
        self.db.commit()

    def _mark_sent(self, target_date, tenant_id, recipient):
        statement = (
            update(NOTIFICATION_LOG_TABLE)
            .where(
                and_(
                    NOTIFICATION_LOG_TABLE.c.notification_date == target_date,
                    NOTIFICATION_LOG_TABLE.c.tenant_id == tenant_id,
                    NOTIFICATION_LOG_TABLE.c.recipient == recipient,
                )
            )
            .values(status="sent", sent_at=func.now())
        )
        self.db.execute(statement)
        self.db.commit()

    def _get_admin_recipients(self, tenant_id):
        statement = (
            select(UserTenantDB.email)
            .join(UserTenantRoleDB, UserTenantRoleDB.user_tenant_id == UserTenantDB.id)
            .join(RoleDB, RoleDB.id == UserTenantRoleDB.role_id)
            .where(
                and_(
                    UserTenantDB.tenant_id == tenant_id,
                    UserTenantDB.status == 1,
                    RoleDB.tenant_id == tenant_id,
                    RoleDB.code.in_(ADMIN_ROLE_CODES),
                    RoleDB.status == 1,
                )
            )
            .distinct()
        )
        rows = self.db.execute(statement).scalars().all()
        return [email.strip() for email in rows if email and email.strip()]
